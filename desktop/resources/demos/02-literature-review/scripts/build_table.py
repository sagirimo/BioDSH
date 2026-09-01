# -*- coding: utf-8 -*-
"""
根据人工精选的 20 篇文献（curated_selection.json）生成 literature_table.xlsx。
- 每个 DOI 通过 Europe PMC 重新核对元数据（题目/一作/年份/期刊/被引数），失败则回退 candidates.json
- 清洗标题中的 HTML 标签（<sup> 等）
- Sheet1: 文献列表；Sheet2: 检索与筛选说明
"""
import html
import json
import os
import re
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE, "analysis_outputs")
EPMC = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
HEADERS = {"User-Agent": "BioDSH-literature-review/1.0"}

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def clean(text):
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", text)      # 去 HTML 标签
    text = html.unescape(text)               # 解实体
    return re.sub(r"\s+", " ", text).strip().rstrip(".")


def get_json(url, params, tries=4, timeout=60):
    last = None
    for i in range(tries):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=timeout)
            if r.status_code == 200:
                return r.json()
            last = RuntimeError("HTTP %s" % r.status_code)
        except Exception as e:  # noqa: BLE001
            last = e
        time.sleep(1.5 * (i + 1))
    raise RuntimeError("lookup failed: %s" % last)


def lookup_doi(doi):
    data = get_json(EPMC, {"query": 'DOI:"%s"' % doi, "format": "json", "resultType": "core"})
    res = data.get("resultList", {}).get("result", [])
    for rec in res:
        if (rec.get("doi") or "").lower() == doi.lower():
            authors = rec.get("authorString", "") or ""
            jinfo = rec.get("journalInfo") or {}
            jtitle = (rec.get("journalTitle") or ""
                      or (jinfo.get("journal") or {}).get("title", "")
                      or (jinfo.get("journal") or {}).get("medlineAbbreviation", ""))
            return {
                "title": clean(rec.get("title", "")),
                "first_author": authors.split(",")[0].strip() if authors else "",
                "year": rec.get("pubYear", ""),
                "journal": jtitle,
                "citations": int(rec.get("citedByCount") or 0),
                "pmid": rec.get("id") if rec.get("source") == "MED" else "",
            }
    return None


def load_candidates():
    p = os.path.join(OUT_DIR, "candidates.json")
    if not os.path.exists(p):
        return {}
    with open(p, encoding="utf-8") as f:
        data = json.load(f)
    out = {}
    for c in data["candidates"]:
        if c.get("doi"):
            out[c["doi"].lower()] = c
    return out


def main():
    with open(os.path.join(BASE, "scripts", "curated_selection.json"), encoding="utf-8") as f:
        curated = json.load(f)["papers"]

    cand = load_candidates()
    rows = []
    problems = []
    for i, item in enumerate(curated, 1):
        doi = item["doi"].lower()
        rec = None
        try:
            rec = lookup_doi(doi)
        except Exception as e:  # noqa: BLE001
            problems.append("%s lookup error: %s" % (doi, e))
        if rec is None and doi in cand:
            c = cand[doi]
            rec = {
                "title": clean(c.get("title", "")),
                "first_author": c.get("first_author", ""),
                "year": c.get("year", ""),
                "journal": c.get("journal", "") or "",
                "citations": c.get("citations") or 0,
                "pmid": c.get("pmid", ""),
            }
        if rec is None:
            problems.append("%s 元数据缺失" % doi)
            rec = {"title": item.get("title", "?"), "first_author": "?", "year": "?",
                   "journal": "?", "citations": 0, "pmid": ""}
        rows.append({
            "no": i,
            "title": rec["title"] or item.get("title", "?"),
            "first_author": rec["first_author"] or item.get("first_author", "?"),
            "year": rec["year"] or item.get("year", "?"),
            "journal": rec["journal"] or item.get("journal", "?"),
            "doi": doi,
            "citations": rec["citations"],
            "conclusion": item["conclusion"],
        })

    # 打印核对
    print("== 核对表 ==")
    for r in rows:
        print("%2d. [%s] %s | %s | %s cites | %s" % (
            r["no"], r["year"], r["first_author"], r["journal"], r["citations"], r["title"][:70]))
    if problems:
        print("== 问题 ==")
        for p in problems:
            print(" -", p)

    # 写 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "文献列表"
    headers = ["序号", "题目", "第一作者", "年份", "期刊", "DOI", "被引次数", "一句话核心结论"]
    ws.append(headers)
    for r in rows:
        ws.append([r["no"], r["title"], r["first_author"], r["year"], r["journal"],
                   r["doi"], r["citations"], r["conclusion"]])

    # 样式
    head_fill = PatternFill("solid", fgColor="2F5B8F")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [6, 62, 14, 8, 26, 30, 10, 66]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        doi_cell = row[5]
        doi_cell.hyperlink = "https://doi.org/" + doi_cell.value
        doi_cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"

    # Sheet2: 检索说明
    ws2 = wb.create_sheet("检索说明")
    log = {}
    try:
        with open(os.path.join(OUT_DIR, "search_log.json"), encoding="utf-8") as f:
            log = json.load(f)
    except Exception:  # noqa: BLE001
        log = {}
    notes = [
        ["检索主题", "肿瘤浸润 T 细胞耗竭（T cell exhaustion）与免疫检查点抑制剂疗效的关系"],
        ["关键词", "T cell exhaustion; PD-1; anti-PD-1 response; tumor-infiltrating lymphocytes; single-cell RNA-seq"],
        ["数据库/接口", "PubMed E-utilities (esearch/esummary) + Europe PMC REST API（Python 调用，无浏览器）"],
        ["时间范围", "2019-01-01 至今（2019–2026）"],
        ["文献类型", "原始研究为主（人类样本优先），辅以少量高被引综述"],
        ["去重方式", "按 PMID/DOI 去重"],
        ["初检总量", "%s 篇（去重后）" % log.get("total_unique", "?"), ],
        ["筛选流程", "按 6 组主题检索式×（相关度/被引两种排序）+ 高被引地标池 合并 → 相关性门槛 → 综合打分（主题面覆盖、标题关键词、被引量、人类研究偏好、综述降权、新近度）→ 取前 32 → 人工精选 20 篇"],
        ["检索式", ""],
    ]
    for n in notes:
        ws2.append(n)
    if log.get("europepmc"):
        ws2.append(["—— Europe PMC 检索明细 ——"])
        for e in log["europepmc"]:
            if "error" in e:
                ws2.append([e.get("facet"), e.get("sort", ""), "ERROR: " + str(e["error"])])
            else:
                ws2.append([e.get("facet"), e.get("sort", ""), "命中 %s，取回 %s" % (e.get("hits"), e.get("returned")),
                            e.get("query", "")])
    if log.get("pubmed"):
        ws2.append(["—— PubMed 检索明细 ——"])
        for e in log["pubmed"]:
            if "error" in e:
                ws2.append([e.get("facet"), "ERROR: " + str(e["error"])])
            else:
                ws2.append([e.get("facet"), "命中 %s，取回 %s" % (e.get("hits"), e.get("returned"))])
    ws2.append(["检索完成时间", log.get("finished", "?")])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 42
    ws2.column_dimensions["D"].width = 90
    for row in ws2.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    out_path = os.path.join(BASE, "literature_table.xlsx")
    wb.save(out_path)
    print("saved:", out_path)


if __name__ == "__main__":
    main()
