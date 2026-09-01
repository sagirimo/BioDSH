# -*- coding: utf-8 -*-
"""
从 NCBI Gene / UniProt / Ensembl 抓取基因注释，输出 gene_annotations.xlsx
用法: python scripts/fetch_annotations.py
"""
import csv
import io
import os
import time
import xml.etree.ElementTree as ET

import pandas as pd
import requests

HDR = {"User-Agent": "BioDSH-demo/03 (T-cell-exhaustion gene annotation; contact: research)"}
NCBI = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
UNIPROT = "https://rest.uniprot.org/uniprotkb"
ENSEMBL = "https://rest.ensembl.org"

WORKDIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def ncbi_gene(symbol):
    """NCBI Gene: 返回 (gene_id, full_name, summary) 或 None"""
    term = f'{symbol}[Gene Name] AND Homo sapiens[Organism]'
    r = requests.get(f"{NCBI}/esearch.fcgi",
                     params={"db": "gene", "term": term, "retmode": "json"},
                     headers=HDR, timeout=30)
    r.raise_for_status()
    idlist = r.json().get("esearchresult", {}).get("idlist", [])
    if not idlist:
        return None, "NCBI: 未找到该基因"
    gid = idlist[0]
    time.sleep(0.4)
    e = requests.get(f"{NCBI}/efetch.fcgi",
                     params={"db": "gene", "id": gid, "retmode": "xml"},
                     headers=HDR, timeout=30)
    e.raise_for_status()
    root = ET.fromstring(e.text)
    gene = root.find(".//Entrezgene")
    if gene is None:
        return None, "NCBI: 解析 XML 失败"
    gene_id = gene.findtext(".//Gene-track_geneid")
    name = gene.findtext(".//Gene-ref_desc") or gene.findtext(".//Gene-ref_locus")
    summary = ""
    comments = gene.find("Entrezgene_comments")
    if comments is not None:
        for gc in comments.findall("Gene-commentary"):
            t = gc.find("Gene-commentary_type")
            if t is not None and t.get("value") == "summary":
                summary = (gc.findtext("Gene-commentary_text") or "").strip()
                break
    return (gene_id, name, summary), None


def uniprot_gene(symbol):
    """UniProt: 返回 (uniprot_id, protein_name, function, ensembl_xref) 或 None"""
    def query(reviewed):
        q = f"gene_exact:{symbol} AND organism_id:9606" + (" AND reviewed:true" if reviewed else "")
        r = requests.get(f"{UNIPROT}/search",
                         params={"query": q,
                                 "fields": "accession,protein_name,cc_function,xref_ensembl",
                                 "format": "tsv", "size": 1},
                         headers=HDR, timeout=30)
        r.raise_for_status()
        rows = list(csv.DictReader(io.StringIO(r.text), delimiter="\t"))
        return rows[0] if rows else None

    row = query(True)
    if row is None:
        row = query(False)
    if row is None:
        return None, "UniProt: 未找到"
    acc = row.get("Entry", "") or row.get("Accession", "")
    acc = acc.strip()
    pname = (row.get("Protein names") or "").strip().split(" (")[0].strip()
    func = (row.get("Function [CC]") or "").strip()
    ens_xref = (row.get("Ensembl") or row.get("Cross-reference (Ensembl)") or "").strip()
    return (acc, pname, func, ens_xref), None


def ensembl_gene(symbol):
    """Ensembl REST: 返回 (ensembl_id, description) 或 None"""
    r = requests.get(f"{ENSEMBL}/lookup/symbol/homo_sapiens/{symbol}",
                     headers={**HDR, "Accept": "application/json"}, timeout=30)
    if r.status_code == 404:
        return None, "Ensembl: 未找到该符号"
    r.raise_for_status()
    j = r.json()
    return (j.get("id"), j.get("description")), None


def main():
    with open(os.path.join(WORKDIR, "genes.txt"), encoding="utf-8") as f:
        genes = [ln.strip() for ln in f if ln.strip()]

    records, notes = [], []
    for i, sym in enumerate(genes, 1):
        row = {"基因符号": sym}
        note = []

        try:
            n, err = ncbi_gene(sym)
            if n:
                row["NCBI Gene ID"] = n[0]
                row["官方全名"] = n[1]
                ncbi_summary = n[2]
                note.append("NCBI: ok")
            else:
                ncbi_summary = ""
                note.append(err or "NCBI: 未找到")
        except Exception as ex:
            ncbi_summary = ""
            note.append(f"NCBI: 失败({type(ex).__name__})")

        try:
            u, err = uniprot_gene(sym)
            if u:
                row["UniProt 编号"] = u[0]
                row.setdefault("官方全名", u[1])
                row["功能简介"] = u[2]
                note.append("UniProt: ok")
            else:
                row.setdefault("功能简介", ncbi_summary)
                note.append(err or "UniProt: 未找到")
        except Exception as ex:
            row.setdefault("功能简介", ncbi_summary)
            note.append(f"UniProt: 失败({type(ex).__name__})")

        try:
            e, err = ensembl_gene(sym)
            if e:
                row["Ensembl ID"] = e[0]
                note.append("Ensembl: ok")
            else:
                note.append(err or "Ensembl: 未找到")
        except Exception as ex:
            note.append(f"Ensembl: 失败({type(ex).__name__})")

        if "官方全名" not in row:
            row["官方全名"] = ""
        if "功能简介" not in row:
            row["功能简介"] = ""
        row["状态"] = "; ".join(note)
        records.append(row)
        notes.append(" | ".join(note))
        print(f"[{i}/{len(genes)}] {sym}: " + " | ".join(note))
        time.sleep(0.4)

    df = pd.DataFrame(records)
    df = df[["基因符号", "官方全名", "功能简介", "Ensembl ID", "UniProt 编号", "NCBI Gene ID", "状态"]]

    out = os.path.join(WORKDIR, "gene_annotations.xlsx")
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="GeneAnnotations", index=False)
        ws = xw.sheets["GeneAnnotations"]
        from openpyxl.styles import Alignment, Font, PatternFill
        widths = {"A": 12, "B": 46, "C": 80, "D": 20, "E": 14, "F": 14, "G": 40}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5496")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r in ws.iter_rows(min_row=2):
            for c in r:
                if c.column_letter in ("B", "C", "G"):
                    c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"\n已保存: {out}  ({len(df)} 个基因)")


if __name__ == "__main__":
    main()
